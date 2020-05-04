import openpyxl as xl
import os
import zlib
from openpyxl.styles import Font, colors, Alignment

lst = ['A', 'B', 'C', 'D', 'E', 'F']
font1= Font(name=u'宋体', size=14, italic=False, color=colors.BLACK, bold=True)
font2= Font(name='Calibri', size=10, italic=False, color=colors.BLACK, bold=False)
align1 = Alignment(horizontal='center', vertical='center')
align2 = Alignment(horizontal='left', vertical='center')

def delxl(name):
    os.remove(name)
def newxl(name):
    wb = xl.Workbook()
    return wb
def is_exist(name):
    return os.path.isfile(name)
def initws(ws):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 20
    ws.row_dimensions[1].height = 20
    ws.append(['序号', '内容', '种子地址', '网页地址', '校验'])
    for i in range(0, 5):
        index = lst[i] + '1'
        ws[index].font = font1
        ws[index].alignment = align1
def openxl(name):
    if is_exist(name):
        wb = xl.load_workbook(name)
        return wb
    else:
        wb = newxl(name)
        ws = wb.active
        initws(ws)
        return wb
def itemAdd(name, addr, page_addr, ws):
    crc = zlib.crc32(bytes(addr, encoding='utf-8'))
    row = [ws.max_row, name, addr, page_addr, crc]
    if ws.max_row == 1:
        ws.append(row)
    else:
        i = 0
        for cell in list(ws.columns)[4]:
            i = i + 1
            if i == 1:
                continue
            if cell.value == crc:
                return
        ws.append(row)
    for i in range(0, 5):
        index = lst[i] + str(ws.max_row)
        ws[index].font = font2
        ws[index].alignment = align2
def savexl(wb, name):
    wb.save(name)
def seedAdd(name, addr, page_addr):
    print(addr)
    wb = openxl("F:\pyfile\种子合集.xlsx")
    ws = wb.active
    ws.title = "国产原创区"
    itemAdd(name, addr, page_addr, ws)
    savexl(wb, "F:\pyfile\种子合集.xlsx")



